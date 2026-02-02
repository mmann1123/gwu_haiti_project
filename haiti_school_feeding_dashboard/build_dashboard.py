#!/usr/bin/env python3
"""
Haiti School Feeding Program Dashboard Builder

Creates an Excel dashboard with pre-calculated data and charts.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import os

# File paths
SOURCE_FILE = "Summits_MEL_Database_December_2025.xlsx"
OUTPUT_FILE = "Haiti_School_Feeding_Dashboard.xlsx"

# Styles
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=12)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def excel_date_to_str(val):
    """Convert Excel serial date to string."""
    if pd.isna(val):
        return ""
    try:
        if isinstance(val, (int, float)):
            date = datetime(1899, 12, 30) + timedelta(days=int(val))
            return date.strftime('%b %Y')
        return str(val)
    except:
        return str(val)


def format_header_row(ws, row, start_col=1, end_col=None):
    """Apply header formatting."""
    if end_col is None:
        end_col = ws.max_column
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def load_data(filepath):
    """Load and standardize all data."""
    print(f"Loading data from {filepath}...")

    # Load sheets
    presence = pd.read_excel(filepath, sheet_name='Présence', header=0, skiprows=[1])
    feeding = pd.read_excel(filepath, sheet_name="Taux d'alimentation", header=0)
    schools = pd.read_excel(filepath, sheet_name="Info sur les écoles", header=1)
    headcount = pd.read_excel(filepath, sheet_name="Comptage Physique", header=0)

    # Standardize presence column names
    pres_cols = {
        'Le mois': 'month',
        "Nom de l'établissement": 'school_name',
        "Taille de l'école`": 'school_size',
        "ID de l'établissement": 'school_id',
        'Commune': 'commune',
        'Departement': 'department',
        'Supervisor': 'supervisor',
        'Le taux de présence': 'attendance_rate',
        'Le total Effectif': 'enrollment',
        '% de variation': 'variation_pct',
        'La catégorie de variation': 'variation_category',
        'La raison de la variation': 'variation_reason'
    }

    # Map columns that exist
    presence_renamed = {}
    for old, new in pres_cols.items():
        for col in presence.columns:
            if old.lower().replace("'", "").replace("`", "") in col.lower().replace("'", "").replace("`", ""):
                presence_renamed[col] = new
                break
    presence = presence.rename(columns=presence_renamed)

    # Standardize feeding column names
    feed_cols = {
        'Semaine commencant': 'week_start',
        'Le mois': 'month',
        "Nom de l'établissement": 'school_name',
        "ID de l'établissement": 'school_id',
        'Commune': 'commune',
        'Departement': 'department',
        'Supervisor': 'supervisor',
        "Le nombre de jours d'alimentation prévu": 'days_planned',
        "Le nombre réel de jours d'alimentation": 'days_fed',
        "Le taux d'alimentation": 'feeding_rate',
        "1. La catégorie de non-alimentation": 'nonfeeding_cat1',
        "2. La catégorie de non-alimentation": 'nonfeeding_cat2',
        "3. La catégorie de non-alimentation": 'nonfeeding_cat3',
        "4. La catégorie de non-alimentation": 'nonfeeding_cat4',
        "5. La catégorie de non-alimentation": 'nonfeeding_cat5'
    }

    feeding_renamed = {}
    for old, new in feed_cols.items():
        for col in feeding.columns:
            if old.lower().replace("'", "").replace("'", "") in col.lower().replace("'", "").replace("'", ""):
                feeding_renamed[col] = new
                break
    feeding = feeding.rename(columns=feeding_renamed)

    # Standardize schools column names
    school_cols = {
        "Nom de l'Etablissement": 'school_name',
        "ID de l'Etablissement": 'school_id',
        'Commune': 'commune',
        'Departement': 'department',
        'Supervisor': 'supervisor',
        'Grades served': 'grades',
        'School ownership': 'ownership'
    }

    schools_renamed = {}
    for old, new in school_cols.items():
        for col in schools.columns:
            if old.lower().replace("'", "").replace("'", "") in col.lower().replace("'", "").replace("'", ""):
                schools_renamed[col] = new
                break
    schools = schools.rename(columns=schools_renamed)

    print(f"  Présence: {len(presence)} rows, columns: {list(presence.columns[:8])}")
    print(f"  Feeding: {len(feeding)} rows, columns: {list(feeding.columns[:8])}")
    print(f"  Schools: {len(schools)} rows, columns: {list(schools.columns[:6])}")

    return {
        'presence': presence,
        'feeding': feeding,
        'schools': schools,
        'headcount': headcount
    }


def calculate_metrics(data):
    """Calculate all dashboard metrics."""
    presence = data['presence']
    feeding = data['feeding']
    schools = data['schools']

    metrics = {}

    # Latest month
    if 'month' in presence.columns:
        latest_month = presence['month'].max()
        metrics['latest_month'] = latest_month
        metrics['latest_month_str'] = excel_date_to_str(latest_month)
        pres_current = presence[presence['month'] == latest_month]
    else:
        pres_current = presence

    # KPIs
    metrics['total_schools'] = pres_current['school_id'].nunique() if 'school_id' in pres_current.columns else 0

    if 'attendance_rate' in pres_current.columns:
        metrics['avg_attendance'] = pres_current['attendance_rate'].mean()
        metrics['schools_below_80'] = (pres_current.groupby('school_id')['attendance_rate'].mean() < 0.8).sum() if 'school_id' in pres_current.columns else 0
    else:
        metrics['avg_attendance'] = 0
        metrics['schools_below_80'] = 0

    if 'feeding_rate' in feeding.columns:
        metrics['avg_feeding'] = feeding['feeding_rate'].mean()
    else:
        metrics['avg_feeding'] = 0

    # Monthly trends (attendance)
    if 'month' in presence.columns and 'attendance_rate' in presence.columns:
        att_trend = presence.groupby('month').agg({
            'attendance_rate': 'mean',
            'school_id': 'nunique'
        }).reset_index()
        att_trend.columns = ['month', 'avg_attendance', 'school_count']
        att_trend = att_trend.sort_values('month', ascending=False).head(6)
        metrics['attendance_trend'] = att_trend
    else:
        metrics['attendance_trend'] = pd.DataFrame()

    # Monthly trends (feeding)
    if 'month' in feeding.columns and 'feeding_rate' in feeding.columns:
        feed_trend = feeding.groupby('month').agg({
            'feeding_rate': 'mean',
            'days_planned': 'sum',
            'days_fed': 'sum'
        }).reset_index() if 'days_planned' in feeding.columns else feeding.groupby('month')['feeding_rate'].mean().reset_index()
        metrics['feeding_trend'] = feed_trend.head(6)
    else:
        metrics['feeding_trend'] = pd.DataFrame()

    # Commune stats
    if 'commune' in pres_current.columns and 'attendance_rate' in pres_current.columns:
        commune_stats = pres_current.groupby('commune').agg({
            'attendance_rate': 'mean',
            'school_id': 'nunique'
        }).reset_index()
        commune_stats.columns = ['commune', 'avg_attendance', 'school_count']
        commune_stats = commune_stats.sort_values('avg_attendance', ascending=False)
        metrics['commune_stats'] = commune_stats
    else:
        metrics['commune_stats'] = pd.DataFrame()

    # Schools needing attention
    if 'school_name' in pres_current.columns and 'attendance_rate' in pres_current.columns:
        school_stats = pres_current.groupby(['school_id', 'school_name', 'commune', 'supervisor']).agg({
            'attendance_rate': 'mean'
        }).reset_index()
        alerts = school_stats[school_stats['attendance_rate'] < 0.8].sort_values('attendance_rate')
        metrics['alerts'] = alerts
    else:
        metrics['alerts'] = pd.DataFrame()

    # Variation reasons
    if 'variation_category' in pres_current.columns:
        var_reasons = pres_current['variation_category'].value_counts().reset_index()
        var_reasons.columns = ['category', 'count']
        metrics['variation_reasons'] = var_reasons
    else:
        metrics['variation_reasons'] = pd.DataFrame()

    # Non-feeding reasons
    nonfeeding_cols = [c for c in feeding.columns if 'nonfeeding_cat' in c]
    all_reasons = []
    for col in nonfeeding_cols:
        all_reasons.extend(feeding[col].dropna().tolist())
    if all_reasons:
        reason_counts = pd.Series(all_reasons).value_counts().reset_index()
        reason_counts.columns = ['category', 'count']
        metrics['nonfeeding_reasons'] = reason_counts
    else:
        metrics['nonfeeding_reasons'] = pd.DataFrame()

    # Supervisor stats
    if 'supervisor' in pres_current.columns and 'attendance_rate' in pres_current.columns:
        sup_stats = pres_current.groupby('supervisor').agg({
            'attendance_rate': 'mean',
            'school_id': 'nunique'
        }).reset_index()
        sup_stats.columns = ['supervisor', 'avg_attendance', 'school_count']

        # Count schools below 80%
        schools_below = pres_current[pres_current['attendance_rate'] < 0.8].groupby('supervisor')['school_id'].nunique().reset_index()
        schools_below.columns = ['supervisor', 'schools_below_80']

        sup_stats = sup_stats.merge(schools_below, on='supervisor', how='left')
        sup_stats['schools_below_80'] = sup_stats['schools_below_80'].fillna(0).astype(int)
        sup_stats = sup_stats.sort_values('avg_attendance', ascending=False)
        metrics['supervisor_stats'] = sup_stats
    else:
        metrics['supervisor_stats'] = pd.DataFrame()

    # Schools list
    metrics['schools'] = schools

    return metrics


def create_executive_summary(wb, metrics):
    """Create Executive Summary sheet."""
    print("Creating Executive Summary...")
    ws = wb.create_sheet("Executive Summary")

    # Title
    ws['A1'] = "Haiti School Feeding Program - Executive Summary"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')

    # Current month
    ws['A3'] = "Current Month:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = metrics.get('latest_month_str', 'All Data')

    ws['D3'] = "Attendance Target:"
    ws['E3'] = 0.9
    ws['E3'].number_format = '0%'
    ws['D4'] = "Feeding Target:"
    ws['E4'] = 0.9
    ws['E4'].number_format = '0%'

    # KPIs Section
    ws['A6'] = "Key Performance Indicators"
    ws['A6'].font = SUBTITLE_FONT

    headers = ['Metric', 'Value', 'Target', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=7, column=col, value=h)
    format_header_row(ws, 7, 1, 4)

    # KPI data
    kpis = [
        ('Total Schools Active', metrics['total_schools'], '-', '✓'),
        ('Average Attendance Rate', metrics['avg_attendance'], 0.9,
         '✓' if metrics['avg_attendance'] >= 0.9 else '⚠' if metrics['avg_attendance'] >= 0.8 else '✗'),
        ('Average Feeding Rate', metrics['avg_feeding'], 0.9,
         '✓' if metrics['avg_feeding'] >= 0.9 else '⚠' if metrics['avg_feeding'] >= 0.8 else '✗'),
        ('Schools Below 80%', metrics['schools_below_80'], 0,
         '✓' if metrics['schools_below_80'] == 0 else '⚠')
    ]

    for i, (metric, value, target, status) in enumerate(kpis, 8):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value=value)
        if isinstance(value, float) and value <= 1:
            ws.cell(row=i, column=2).number_format = '0.0%'
        ws.cell(row=i, column=3, value=target)
        if isinstance(target, float) and target <= 1:
            ws.cell(row=i, column=3).number_format = '0%'
        ws.cell(row=i, column=4, value=status)

        # Color coding
        if i in [9, 10]:  # Attendance and feeding rows
            if value >= 0.8:
                ws.cell(row=i, column=2).fill = GREEN_FILL
            elif value >= 0.7:
                ws.cell(row=i, column=2).fill = YELLOW_FILL
            else:
                ws.cell(row=i, column=2).fill = RED_FILL

    # Alerts Section
    ws['A14'] = "Schools Needing Attention (Attendance < 80%)"
    ws['A14'].font = SUBTITLE_FONT

    headers = ['School Name', 'Commune', 'Supervisor', 'Attendance']
    for col, h in enumerate(headers, 1):
        ws.cell(row=15, column=col, value=h)
    format_header_row(ws, 15, 1, 4)

    alerts = metrics['alerts']
    if len(alerts) > 0:
        for i, (_, row) in enumerate(alerts.head(10).iterrows(), 16):
            ws.cell(row=i, column=1, value=row.get('school_name', ''))
            ws.cell(row=i, column=2, value=row.get('commune', ''))
            ws.cell(row=i, column=3, value=row.get('supervisor', ''))
            att = row.get('attendance_rate', 0)
            ws.cell(row=i, column=4, value=att)
            ws.cell(row=i, column=4).number_format = '0.0%'
            ws.cell(row=i, column=4).fill = RED_FILL if att < 0.7 else YELLOW_FILL
    else:
        ws['A16'] = "No schools below 80% threshold"

    # Commune Comparison Section
    ws['F6'] = "Attendance by Commune"
    ws['F6'].font = SUBTITLE_FONT

    headers = ['Commune', 'Avg Attendance', 'Schools']
    for col, h in enumerate(headers, 6):
        ws.cell(row=7, column=col, value=h)
    format_header_row(ws, 7, 6, 8)

    commune_stats = metrics['commune_stats']
    for i, (_, row) in enumerate(commune_stats.head(12).iterrows(), 8):
        ws.cell(row=i, column=6, value=row['commune'])
        ws.cell(row=i, column=7, value=row['avg_attendance'])
        ws.cell(row=i, column=7).number_format = '0.0%'
        ws.cell(row=i, column=8, value=row['school_count'])

    # Add bar chart
    if len(commune_stats) > 0:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Attendance by Commune"
        data_end = min(8 + len(commune_stats), 19)
        chart_data = Reference(ws, min_col=7, min_row=7, max_row=data_end)
        cats = Reference(ws, min_col=6, min_row=8, max_row=data_end)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 12
        chart.height = 8
        ws.add_chart(chart, "J6")

    # Trend Section
    ws['A28'] = "6-Month Attendance Trend"
    ws['A28'].font = SUBTITLE_FONT

    headers = ['Month', 'Avg Attendance', 'Schools']
    for col, h in enumerate(headers, 1):
        ws.cell(row=29, column=col, value=h)
    format_header_row(ws, 29, 1, 3)

    att_trend = metrics['attendance_trend']
    for i, (_, row) in enumerate(att_trend.iterrows(), 30):
        ws.cell(row=i, column=1, value=excel_date_to_str(row['month']))
        ws.cell(row=i, column=2, value=row['avg_attendance'])
        ws.cell(row=i, column=2).number_format = '0.0%'
        ws.cell(row=i, column=3, value=row['school_count'])

    # Add line chart
    if len(att_trend) > 1:
        chart = LineChart()
        chart.style = 10
        chart.title = "Attendance Trend"
        data_end = 29 + len(att_trend)
        chart_data = Reference(ws, min_col=2, min_row=29, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=30, max_row=data_end)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 10
        chart.height = 6
        ws.add_chart(chart, "E28")

    # Column widths
    for col, width in enumerate([25, 15, 12, 10, 5, 20, 15, 10], 1):
        ws.column_dimensions[chr(64 + col)].width = width


def create_attendance_analysis(wb, metrics, data):
    """Create Attendance Analysis sheet."""
    print("Creating Attendance Analysis...")
    ws = wb.create_sheet("Attendance Analysis")

    ws['A1'] = "Attendance Analysis"
    ws['A1'].font = TITLE_FONT

    # Trend section
    ws['A4'] = "Monthly Attendance Trend"
    ws['A4'].font = SUBTITLE_FONT

    headers = ['Month', 'Avg Attendance', 'School Count']
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    format_header_row(ws, 5, 1, 3)

    att_trend = metrics['attendance_trend']
    for i, (_, row) in enumerate(att_trend.iterrows(), 6):
        ws.cell(row=i, column=1, value=excel_date_to_str(row['month']))
        ws.cell(row=i, column=2, value=row['avg_attendance'])
        ws.cell(row=i, column=2).number_format = '0.0%'
        ws.cell(row=i, column=3, value=row['school_count'])

    # Add chart
    if len(att_trend) > 1:
        chart = LineChart()
        chart.style = 10
        chart.title = "6-Month Attendance Trend"
        data_end = 5 + len(att_trend)
        chart_data = Reference(ws, min_col=2, min_row=5, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=6, max_row=data_end)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 12
        chart.height = 7
        ws.add_chart(chart, "E4")

    # Variation Reasons
    ws['A15'] = "Attendance Variation Reasons"
    ws['A15'].font = SUBTITLE_FONT

    headers = ['Category', 'Count']
    for col, h in enumerate(headers, 1):
        ws.cell(row=16, column=col, value=h)
    format_header_row(ws, 16, 1, 2)

    var_reasons = metrics['variation_reasons']
    for i, (_, row) in enumerate(var_reasons.head(10).iterrows(), 17):
        ws.cell(row=i, column=1, value=row['category'])
        ws.cell(row=i, column=2, value=row['count'])

    if len(var_reasons) == 0:
        ws['A17'] = "No variation reasons recorded"

    # Pie chart
    if len(var_reasons) > 0:
        chart = PieChart()
        chart.title = "Variation Reasons"
        data_end = 16 + len(var_reasons.head(8))
        chart_data = Reference(ws, min_col=2, min_row=16, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=17, max_row=data_end)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 10
        chart.height = 7
        ws.add_chart(chart, "E15")

    # Schools with drops
    ws['A30'] = "Schools with Attendance Drop > 5%"
    ws['A30'].font = SUBTITLE_FONT

    presence = data['presence']
    if 'variation_pct' in presence.columns:
        drops = presence[presence['variation_pct'] < -0.05].copy()
        drops = drops.sort_values('variation_pct').head(15)

        headers = ['School', 'Commune', 'Current Rate', 'Change']
        for col, h in enumerate(headers, 1):
            ws.cell(row=31, column=col, value=h)
        format_header_row(ws, 31, 1, 4)

        for i, (_, row) in enumerate(drops.iterrows(), 32):
            ws.cell(row=i, column=1, value=row.get('school_name', ''))
            ws.cell(row=i, column=2, value=row.get('commune', ''))
            ws.cell(row=i, column=3, value=row.get('attendance_rate', 0))
            ws.cell(row=i, column=3).number_format = '0.0%'
            ws.cell(row=i, column=4, value=row.get('variation_pct', 0))
            ws.cell(row=i, column=4).number_format = '+0.0%;-0.0%'
    else:
        ws['A32'] = "Variation data not available"

    for col, width in enumerate([40, 18, 15, 12], 1):
        ws.column_dimensions[chr(64 + col)].width = width


def create_feeding_analysis(wb, metrics, data):
    """Create Feeding Rate Analysis sheet."""
    print("Creating Feeding Rate Analysis...")
    ws = wb.create_sheet("Feeding Rate Analysis")

    ws['A1'] = "Feeding Rate Analysis"
    ws['A1'].font = TITLE_FONT

    feeding = data['feeding']

    # Feeding trend
    ws['A4'] = "Feeding Rate Summary"
    ws['A4'].font = SUBTITLE_FONT

    headers = ['Month', 'Avg Feeding Rate', 'Days Planned', 'Days Fed', 'Days Missed']
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    format_header_row(ws, 5, 1, 5)

    if 'month' in feeding.columns:
        feed_by_month = feeding.groupby('month').agg({
            'feeding_rate': 'mean',
            'days_planned': 'sum' if 'days_planned' in feeding.columns else 'count',
            'days_fed': 'sum' if 'days_fed' in feeding.columns else 'count'
        }).reset_index()
        feed_by_month = feed_by_month.sort_values('month', ascending=False).head(6)

        for i, (_, row) in enumerate(feed_by_month.iterrows(), 6):
            ws.cell(row=i, column=1, value=str(row['month']))
            ws.cell(row=i, column=2, value=row['feeding_rate'] if pd.notna(row['feeding_rate']) else 0)
            ws.cell(row=i, column=2).number_format = '0.0%'
            planned = row.get('days_planned', 0) or 0
            fed = row.get('days_fed', 0) or 0
            ws.cell(row=i, column=3, value=planned)
            ws.cell(row=i, column=4, value=fed)
            ws.cell(row=i, column=5, value=planned - fed)

            # Color code feeding rate
            rate = row['feeding_rate'] if pd.notna(row['feeding_rate']) else 0
            if rate >= 0.8:
                ws.cell(row=i, column=2).fill = GREEN_FILL
            elif rate >= 0.7:
                ws.cell(row=i, column=2).fill = YELLOW_FILL
            else:
                ws.cell(row=i, column=2).fill = RED_FILL

        # Stacked bar chart
        if len(feed_by_month) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Days Fed vs Missed"
            chart.grouping = "stacked"
            data_end = 5 + len(feed_by_month)
            chart_data = Reference(ws, min_col=4, min_row=5, max_col=5, max_row=data_end)
            cats = Reference(ws, min_col=1, min_row=6, max_row=data_end)
            chart.add_data(chart_data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 12
            chart.height = 7
            ws.add_chart(chart, "G4")

    # Non-feeding reasons
    ws['A15'] = "Non-Feeding Reasons"
    ws['A15'].font = SUBTITLE_FONT

    headers = ['Category', 'Count']
    for col, h in enumerate(headers, 1):
        ws.cell(row=16, column=col, value=h)
    format_header_row(ws, 16, 1, 2)

    nf_reasons = metrics['nonfeeding_reasons']
    for i, (_, row) in enumerate(nf_reasons.head(10).iterrows(), 17):
        ws.cell(row=i, column=1, value=row['category'])
        ws.cell(row=i, column=2, value=row['count'])

    if len(nf_reasons) == 0:
        ws['A17'] = "No non-feeding reasons recorded"

    # Bar chart for reasons
    if len(nf_reasons) > 0:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Non-Feeding Reasons"
        data_end = 16 + len(nf_reasons.head(8))
        chart_data = Reference(ws, min_col=2, min_row=16, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=17, max_row=data_end)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 10
        chart.height = 6
        ws.add_chart(chart, "D15")

    # Detailed table
    ws['A30'] = "Recent Feeding Data"
    ws['A30'].font = SUBTITLE_FONT

    headers = ['School', 'Commune', 'Week', 'Days Planned', 'Days Fed', 'Rate', 'Non-Feed Reason']
    for col, h in enumerate(headers, 1):
        ws.cell(row=31, column=col, value=h)
    format_header_row(ws, 31, 1, 7)

    recent = feeding.head(20)
    for i, (_, row) in enumerate(recent.iterrows(), 32):
        ws.cell(row=i, column=1, value=row.get('school_name', ''))
        ws.cell(row=i, column=2, value=row.get('commune', ''))
        ws.cell(row=i, column=3, value=excel_date_to_str(row.get('week_start', '')))
        ws.cell(row=i, column=4, value=row.get('days_planned', 0))
        ws.cell(row=i, column=5, value=row.get('days_fed', 0))
        rate = row.get('feeding_rate', 0)
        ws.cell(row=i, column=6, value=rate if pd.notna(rate) else 0)
        ws.cell(row=i, column=6).number_format = '0.0%'
        ws.cell(row=i, column=7, value=row.get('nonfeeding_cat1', ''))

    for col, width in enumerate([40, 15, 12, 12, 10, 10, 20], 1):
        ws.column_dimensions[chr(64 + col)].width = width


def create_school_detail(wb, metrics, data):
    """Create School Detail sheet."""
    print("Creating School Detail...")
    ws = wb.create_sheet("School Detail")

    ws['A1'] = "School Detail View"
    ws['A1'].font = TITLE_FONT

    ws['A3'] = "Instructions: This sheet shows a sample school. In Google Sheets, add a dropdown to select different schools."
    ws['A3'].font = Font(italic=True)

    # School Profile
    ws['A5'] = "School Profile"
    ws['A5'].font = SUBTITLE_FONT
    ws['A5'].fill = LIGHT_BLUE_FILL
    ws.merge_cells('A5:D5')

    schools = metrics['schools']
    if len(schools) > 0:
        sample = schools.iloc[0]

        profile_data = [
            ('School ID:', sample.get('school_id', '')),
            ('School Name:', sample.get('school_name', '')),
            ('Commune:', sample.get('commune', '')),
            ('Department:', sample.get('department', '')),
            ('Supervisor:', sample.get('supervisor', '')),
            ('Grades:', sample.get('grades', '')),
            ('Type:', sample.get('ownership', ''))
        ]

        for i, (label, value) in enumerate(profile_data, 6):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=1).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)

    # All schools reference
    ws['A16'] = "All Schools Reference"
    ws['A16'].font = SUBTITLE_FONT

    headers = ['School Name', 'ID', 'Commune', 'Department', 'Supervisor']
    for col, h in enumerate(headers, 1):
        ws.cell(row=17, column=col, value=h)
    format_header_row(ws, 17, 1, 5)

    for i, (_, row) in enumerate(schools.head(50).iterrows(), 18):
        ws.cell(row=i, column=1, value=row.get('school_name', ''))
        ws.cell(row=i, column=2, value=row.get('school_id', ''))
        ws.cell(row=i, column=3, value=row.get('commune', ''))
        ws.cell(row=i, column=4, value=row.get('department', ''))
        ws.cell(row=i, column=5, value=row.get('supervisor', ''))

    for col, width in enumerate([45, 20, 15, 12, 25], 1):
        ws.column_dimensions[chr(64 + col)].width = width


def create_supervisor_performance(wb, metrics):
    """Create Supervisor Performance sheet."""
    print("Creating Supervisor Performance...")
    ws = wb.create_sheet("Supervisor Performance")

    ws['A1'] = "Supervisor Performance"
    ws['A1'].font = TITLE_FONT

    ws['A4'] = "Supervisor Ranking"
    ws['A4'].font = SUBTITLE_FONT

    headers = ['Rank', 'Supervisor', 'Schools', 'Avg Attendance', 'Schools < 80%', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    format_header_row(ws, 5, 1, 6)

    sup_stats = metrics['supervisor_stats']
    for i, (_, row) in enumerate(sup_stats.iterrows(), 6):
        ws.cell(row=i, column=1, value=i - 5)
        ws.cell(row=i, column=2, value=row['supervisor'])
        ws.cell(row=i, column=3, value=row['school_count'])
        ws.cell(row=i, column=4, value=row['avg_attendance'])
        ws.cell(row=i, column=4).number_format = '0.0%'
        ws.cell(row=i, column=5, value=row['schools_below_80'])

        # Status and coloring
        below = row['schools_below_80']
        if below >= 3:
            status = "⚠ Critical"
            ws.cell(row=i, column=6).fill = RED_FILL
        elif below >= 2:
            status = "⚠ Attention"
            ws.cell(row=i, column=6).fill = YELLOW_FILL
        else:
            status = "✓ OK"
            ws.cell(row=i, column=6).fill = GREEN_FILL
        ws.cell(row=i, column=6, value=status)

        # Color attendance
        att = row['avg_attendance']
        if att >= 0.8:
            ws.cell(row=i, column=4).fill = GREEN_FILL
        elif att >= 0.7:
            ws.cell(row=i, column=4).fill = YELLOW_FILL
        else:
            ws.cell(row=i, column=4).fill = RED_FILL

    # Bar chart
    if len(sup_stats) > 0:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Supervisor Performance"
        data_end = 5 + len(sup_stats.head(15))
        chart_data = Reference(ws, min_col=4, min_row=5, max_row=data_end)
        cats = Reference(ws, min_col=2, min_row=6, max_row=data_end)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 12
        chart.height = 10
        ws.add_chart(chart, "H4")

    # Problem flagging
    problem_sups = sup_stats[sup_stats['schools_below_80'] >= 2]
    row_start = max(6 + len(sup_stats), 25)

    ws.cell(row=row_start, column=1, value="Supervisors with 2+ Problem Schools")
    ws.cell(row=row_start, column=1).font = SUBTITLE_FONT

    if len(problem_sups) > 0:
        headers = ['Supervisor', 'Problem Schools', 'Total Schools', 'Problem Rate']
        for col, h in enumerate(headers, 1):
            ws.cell(row=row_start + 1, column=col, value=h)
        format_header_row(ws, row_start + 1, 1, 4)

        for i, (_, row) in enumerate(problem_sups.iterrows(), row_start + 2):
            ws.cell(row=i, column=1, value=row['supervisor'])
            ws.cell(row=i, column=2, value=row['schools_below_80'])
            ws.cell(row=i, column=3, value=row['school_count'])
            rate = row['schools_below_80'] / row['school_count'] if row['school_count'] > 0 else 0
            ws.cell(row=i, column=4, value=rate)
            ws.cell(row=i, column=4).number_format = '0%'
    else:
        ws.cell(row=row_start + 2, column=1, value="No supervisors with 2+ problem schools - Great!")
        ws.cell(row=row_start + 2, column=1).fill = GREEN_FILL

    for col, width in enumerate([8, 30, 10, 15, 15, 12], 1):
        ws.column_dimensions[chr(64 + col)].width = width


def main():
    """Build the dashboard."""
    print("=" * 60)
    print("Haiti School Feeding Program - Dashboard Builder")
    print("=" * 60)

    if not os.path.exists(SOURCE_FILE):
        print(f"ERROR: Source file not found: {SOURCE_FILE}")
        return

    # Load data
    data = load_data(SOURCE_FILE)

    # Calculate metrics
    print("\nCalculating metrics...")
    metrics = calculate_metrics(data)

    print(f"  Latest month: {metrics.get('latest_month_str', 'N/A')}")
    print(f"  Total schools: {metrics['total_schools']}")
    print(f"  Avg attendance: {metrics['avg_attendance']:.1%}")
    print(f"  Avg feeding: {metrics['avg_feeding']:.1%}")

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Create sheets
    create_executive_summary(wb, metrics)
    create_attendance_analysis(wb, metrics, data)
    create_feeding_analysis(wb, metrics, data)
    create_school_detail(wb, metrics, data)
    create_supervisor_performance(wb, metrics)

    # Save
    print(f"\nSaving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)

    print("\n" + "=" * 60)
    print("Dashboard created successfully!")
    print(f"Output: {OUTPUT_FILE}")
    print("\nSheets:")
    print("  1. Executive Summary - KPIs and overview")
    print("  2. Attendance Analysis - Trends and reasons")
    print("  3. Feeding Rate Analysis - Feeding metrics")
    print("  4. School Detail - Individual school view")
    print("  5. Supervisor Performance - Rankings")
    print("\nNext: Open in Excel or upload to Google Sheets")
    print("=" * 60)


if __name__ == "__main__":
    main()
